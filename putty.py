import subprocess, sys, pyautogui, time
import openpyxl
import pdb
import os

print (sys.argv)
ip = '' + sys.argv[1]
username = ''
password = ''
password_en = ''
mac = ''
port = ''
if  not ('.' in ip):
    ip = '10.80.0.' + sys.argv[1]

elif ip.count('.') == 1:
    ip = '10.50.' + sys.argv[1]
else:
    ip = sys.argv[1]

wb = openpyxl.load_workbook(r'C:\Users\yalina\PycharmProjects\console\client.xlsx')
sheet = wb.get_sheet_by_name('Password')
for i in range(1, sheet.max_row+1):
    if ip in sheet.cell(row=i, column=1).value:
        username = sheet.cell(row=i, column=2).value
        password = sheet.cell(row=i, column=3).value
        password_en = sheet.cell(row=i, column=4).value
        break
else:
    username = input('Username:')
    password = input('Password:')
    password_en = input('Enable Password:')

subprocess.Popen("putty {} -l {} -pw {}".format(ip, username,password,password_en))
time.sleep(1)
if not ip.endswith("254"):
    subprocess.Popen(r"C:\Program Files\AutoHotkey\AutoHotKey.exe C:\Users\yalina\PycharmProjects\console\putty_enable.ahk {}".format(password_en))
