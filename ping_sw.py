import os
import openpyxl


wb = openpyxl.load_workbook('sw_list_by_ping.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
#sheet.cell(row=1, column=2).value = "a"

for i in range(1,255):
    sheet.cell(row=i, column=1).value = "10.80.0.{}".format(str(i))
    response = os.system('ping 10.80.0.{} -w 1 -n 1'.format(str(i)))
    if response == 0:
        print('is up!')
        sheet.cell(row=i, column=2).value = "1"








wb.save('sw_list_by_ping.xlsx')

# sheet.cell(row=i, column=1).value = "10.80.0.{}".format(str(i))
# if
# sheet.cell(row=i, column=2).value = "10.80.0.{}".format(str(i))

# s = os.popen('ping 10.80.0.{} -w 1 -n 2'.format(str(i))).read()
# print(s)