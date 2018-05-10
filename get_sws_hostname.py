import paramiko
import time
import pdb
import re
import pyperclip
import logging
import win32com.client as comclt
import subprocess
import sys
import openpyxl



def disable_paging(remote_conn):
    '''Disable paging on a Cisco router'''

    remote_conn.send("terminal length 0\n")
    time.sleep(1)

    # Clear the buffer on the screen
    output = remote_conn.recv(1000)
    return output


def send_command(remote_conn, command, screen):
    remote_conn.send(command + '\n')
    rcv_timeout = 0
    output_all = ''
    output = b''
    while rcv_timeout < 5.9 and not ((output.decode('utf-8')).endswith('>') or
                                         (output.decode('utf-8')).endswith('#')):
        if remote_conn.recv_ready():
            output = remote_conn.recv(4096)
            output_all = output_all + output.decode('utf-8')
        else:
            time.sleep(0.1)
            rcv_timeout += 0.1
    if screen:
        # print(output)
        print(output_all)
    return output_all

ip_list = [
    '10.80.0.10',
    '10.80.0.102',
    '10.80.0.103',
    '10.80.0.104',
    '10.80.0.105',
    '10.80.0.106',
    '10.80.0.107',
    '10.80.0.108',
    '10.80.0.109',
    '10.80.0.11',
    '10.80.0.110',
    '10.80.0.111',
    '10.80.0.112',
    '10.80.0.113',
    '10.80.0.114',
    '10.80.0.115',
    '10.80.0.116',
    '10.80.0.117',
    '10.80.0.118',
    '10.80.0.119',
    '10.80.0.12',
    '10.80.0.120',
    '10.80.0.121',
    '10.80.0.122',
    '10.80.0.123',
    '10.80.0.124',
    '10.80.0.125',
    '10.80.0.126',
    '10.80.0.127',
    '10.80.0.128',
    '10.80.0.129',
    '10.80.0.13',
    '10.80.0.130',
    '10.80.0.131',
    '10.80.0.15',
    '10.80.0.16',
    '10.80.0.17',
    '10.80.0.18',
    '10.80.0.19',
    '10.80.0.2',
    '10.80.0.20',
    '10.80.0.21',
    '10.80.0.22',
    '10.80.0.221',
    '10.80.0.223',
    '10.80.0.224',
    '10.80.0.23',
    '10.80.0.24',
    '10.80.0.25',
    '10.80.0.26',
    '10.80.0.27',
    '10.80.0.28',
    '10.80.0.29',
    '10.80.0.3',
    '10.80.0.30',
    '10.80.0.32',
    '10.80.0.33',
    '10.80.0.34',
    '10.80.0.35',
    '10.80.0.36',
    '10.80.0.37',
    '10.80.0.38',
    '10.80.0.39',
    '10.80.0.4',
    '10.80.0.40',
    '10.80.0.41',
    '10.80.0.42',
    '10.80.0.43',
    '10.80.0.44',
    '10.80.0.45',
    '10.80.0.47',
    '10.80.0.48',
    '10.80.0.49',
    '10.80.0.5',
    '10.80.0.50',
    '10.80.0.51',
    '10.80.0.52',
    '10.80.0.53',
    '10.80.0.55',
    '10.80.0.56',
    '10.80.0.57',
    '10.80.0.58',
    '10.80.0.59',
    '10.80.0.6',
    '10.80.0.61',
    '10.80.0.62',
    '10.80.0.63',
    '10.80.0.64',
    '10.80.0.65',
    '10.80.0.66',
    '10.80.0.67',
    '10.80.0.68',
    '10.80.0.69',
    '10.80.0.7',
    '10.80.0.71',
    '10.80.0.73',
    '10.80.0.79',
    '10.80.0.8',
    '10.80.0.80',
    '10.80.0.81',
    '10.80.0.82',
    '10.80.0.83',
    '10.80.0.84',
    '10.80.0.85',
    '10.80.0.86',
    '10.80.0.87',
    '10.80.0.88',
    '10.80.0.89',
    '10.80.0.9',
    '10.80.0.94',
]
ip_list2 = [
    '10.80.0.2',
    '10.80.0.94',
]
def get_hostname(ip,username,password):
    remote_conn_pre = paramiko.SSHClient()
    remote_conn_pre.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    remote_conn_pre.connect(ip, username=username, password=password, look_for_keys=False, allow_agent=False)
    remote_conn = remote_conn_pre.invoke_shell()

    output = remote_conn.recv(1000)
    sw_hostname = disable_paging(remote_conn).decode("utf-8")
    print(sw_hostname.splitlines()[1])
    remote_conn_pre.close()
    return sw_hostname.splitlines()[1]

if __name__ == '__main__':
    logging.basicConfig(level=logging.INFO)
    logger = logging.getLogger(__name__)
    wb = openpyxl.load_workbook('sw_hostname.xlsx')
    sheet = wb.get_sheet_by_name('Sheet1')
    ip = '10.80.0.2'
    username = 'amerikan'
    password = 'Password'
    print(len(ip_list))
    for i,ip in enumerate(ip_list):
        sheet.cell(row=i + 1, column=1).value = get_hostname(ip, username, password)
        sheet.cell(row=i+1, column=2).value = ip
        sheet.cell(row=i + 1, column=3).value = username
        sheet.cell(row=i + 1, column=4).value = password
        sheet.cell(row=i + 1, column=5).value = 'Amerikan Hastanesi'


    wb.save('sw_hostname.xlsx')
