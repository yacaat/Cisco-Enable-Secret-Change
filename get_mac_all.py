import os
import openpyxl
import re
import threading
from queue import Queue
import time
lock = threading.Lock()



mac_address_list = list()
#wb = openpyxl.load_workbook('get_mac_all_ah.xlsx')
wb = openpyxl.Workbook()
wb.save('get_mac_all_ah.xlsx')
sheet = wb.get_sheet_by_name('Sheet')

ip_list = [
    '10.80.0.2',
    '10.80.0.3',
    '10.80.0.4',
    '10.80.0.5',
    '10.80.0.6',
    '10.80.0.7',
    '10.80.0.8',
    '10.80.0.9',
    '10.80.0.10',
    '10.80.0.11',
    '10.80.0.12',
    '10.80.0.13',
    '10.80.0.15',
    '10.80.0.16',
    '10.80.0.17',
    '10.80.0.18',
    '10.80.0.19',
    '10.80.0.20',
    '10.80.0.21',
    '10.80.0.22',
    '10.80.0.23',
    '10.80.0.24',
    '10.80.0.25',
    '10.80.0.26',
    '10.80.0.27',
    '10.80.0.28',
    '10.80.0.29',
    '10.80.0.30',
    '10.80.0.32',
    '10.80.0.33',
    '10.80.0.34',
    '10.80.0.35',
    '10.80.0.36',
    '10.80.0.37',
    '10.80.0.38',
    '10.80.0.39',
    '10.80.0.40',
    '10.80.0.41',
    '10.80.0.42',
    '10.80.0.43',
    '10.80.0.44',
    '10.80.0.45',
    '10.80.0.47',
    '10.80.0.48',
    '10.80.0.49',
    '10.80.0.50',
    '10.80.0.51',
    '10.80.0.52',
    '10.80.0.53',
    '10.80.0.55',
    '10.80.0.56',
    '10.80.0.57',
    '10.80.0.58',
    '10.80.0.59',
    '10.80.0.61',
    '10.80.0.62',
    '10.80.0.63',
    '10.80.0.64',
    '10.80.0.65',
    '10.80.0.66',
    '10.80.0.67',
    '10.80.0.68',
    '10.80.0.69',
    '10.80.0.71',
    '10.80.0.73',
    '10.80.0.79',
    '10.80.0.80',
    '10.80.0.81',
    '10.80.0.82',
    '10.80.0.83',
    '10.80.0.84',
    '10.80.0.85',
    '10.80.0.86',
    '10.80.0.87',
    '10.80.0.88',
    '10.80.0.89',
    '10.80.0.94',
    '10.80.0.102',
    '10.80.0.103',
    '10.80.0.104',
    '10.80.0.105',
    '10.80.0.106',
    '10.80.0.107',
    '10.80.0.108',
    '10.80.0.109',
    '10.80.0.110',
    '10.80.0.111',
    '10.80.0.112',
    '10.80.0.113',
    '10.80.0.114',
    '10.80.0.115',
    '10.80.0.116',
    '10.80.0.117',
    '10.80.0.118',
    '10.80.0.119',
    '10.80.0.120',
    '10.80.0.121',
    '10.80.0.122',
    '10.80.0.123',
    '10.80.0.124',
    '10.80.0.125',
    '10.80.0.126',
    '10.80.0.127',
    '10.80.0.128',
    '10.80.0.129',
    '10.80.0.130',
    '10.80.0.131',
]
def get_mac_table(ip):

    print('Trying {}'.format(ip))
    mac_address_table = os.popen('plink {} -l amerikan -pw PASSWORD  "sh mac add"'.format(str(ip))).read()

    macs = re.findall(r'(\w{4}\.\w{4}\.\w{4}).*\s(\S+)', mac_address_table)
    for mac, port in macs:
        mac_address_list.append((ip, mac, port))
        #print(mac + '-' + port)
#
# for ip in ip_list:
#      get_mac_table(ip)
def worker():
    while True:
        item = q.get()
        get_mac_table(item)
        q.task_done()

q = Queue()
for i in range(32):
     t = threading.Thread(target=worker)
     t.daemon = True  # thread dies when main thread (only non-daemon thread) exits.
     t.start()

start = time.perf_counter()
for item in ip_list:
    q.put(item)

q.join()

print('time:',time.perf_counter() - start)
i = 1
for ip in mac_address_list:
    if not('Po1' in ip[2]) and not ('CPU' in ip[2]):
        print("{} -- {} -- {} -- {}".format(i,ip[0],ip[1],ip[2]))
        sheet.cell(row=i, column=1).value = ip[0]
        sheet.cell(row=i, column=2).value = ip[1]
        sheet.cell(row=i, column=3).value = ip[2]
        i += 1

wb.save('get_mac_all_ah.xlsx')